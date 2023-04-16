from aiogram import Bot, Dispatcher, executor, types
from openpyxl.reader.excel import load_workbook
import config as nav
global row
row=[0,0]
def botik():
    bufname=''
    BOT_TOKEN='6024152220:AAHt7nUYHdf_fHSCjUHYd9NYL_yg77OrWas'
    bot =Bot(token=BOT_TOKEN)
    dp = Dispatcher(bot)
    @dp.message_handler(commands=['start'])
    async def send_welcome(message: types.Message):
        await bot.send_message(message.from_user.id,"Привет! нажми кнопку записи в меню или кнопку проверку записи, чтобы узнать записан ли ты", reply_markup=nav.otherMenu)
    @dp.message_handler(text=['Ввести тип'])
    async def send_message(message: types.Message):
        await bot.send_message(message.from_user.id, 'введите тип марафона', reply_markup=nav.runMenu)
    @dp.message_handler(text=['Подтвердить участие'])
    async def send_message(message: types.Message):
        print (row)
        if row[0]!=0 and row[1]!=0:
            vvod()
            await bot.send_message(message.from_user.id, 'вы записались')
            rowdestroy()
        else:
            print('не все данные введены')
            await bot.send_message(message.from_user.id, 'не все данные введены')
    @dp.message_handler(text=['Проверить запись'])
    async def send_message(message: types.Message):
        print(row)
        if row[0]!=0 and row[1]!=0:
            if prov()==0:
                await bot.send_message(message.from_user.id, 'вы  записаны')
            else:
                await bot.send_message(message.from_user.id, 'вы не записаны')
            rowdestroy()
        else:
            print('не все данные введены')
            await bot.send_message(message.from_user.id, 'не все данные введены')
    @dp.message_handler(text=['Узнать об участии'])
    async def send_message(message: types.Message):
        await bot.send_message(message.from_user.id, 'введите тип марафона и имя', reply_markup=nav.otherMenu)
    @dp.message_handler()
    async def send_message(message: types.Message):
        if message.text=='Марафон':
            #global type
            type=message.text
            await bot.send_message(message.from_user.id, 'принял', reply_markup=nav.otherMenu)
            rowcreate(type, 1)
        elif message.text=='Полумарафон':
            type = message.text
            await bot.send_message(message.from_user.id, 'принял', reply_markup=nav.otherMenu)
            rowcreate(type, 1)
        elif message.text=='10000 метров':
            type = message.text
            await bot.send_message(message.from_user.id, 'принял', reply_markup=nav.otherMenu)
            rowcreate(type,1)
        elif message.text=='Ввести имя':
            global bufname
            bufname=message.text
            await bot.send_message(message.from_user.id, 'Вводите имя',reply_markup=types.ReplyKeyboardRemove())
        elif bufname=='Ввести имя':
            #global name
            name = message.text
            rowcreate(name,0)
            bufname=''
            await bot.send_message(message.from_user.id, 'принял', reply_markup=nav.otherMenu)
        else:
            await bot.send_message(message.from_user.id, 'выберите что-нибудь')

    executor.start_polling(dp, skip_updates=True)
def vvod():
    fn = r"C:\Users\msiuser\PycharmProjects\лаба2\laba2.xlsx"
    wb = load_workbook(fn)
    ws = wb["Sheet1"]
    ws.append(row)
    wb.save(fn)
    wb.close()
def prov():
    fn = r"C:\Users\msiuser\PycharmProjects\лаба2\laba2.xlsx"
    wb = load_workbook(fn)
    sheet = wb.active
    m_row = sheet.max_row
    for i in range(2, m_row + 1):
        namecell = sheet.cell(row=i, column=1)
        typecell = sheet.cell(row=i, column=2)
        if namecell.value == row[0] and typecell.value == row[1]:
            wb.close()
            return 0
    wb.close()
    return 1
def rowdestroy():
    row[0]=0
    row[1]=0
def rowcreate(a,b):
    row[b]=a

